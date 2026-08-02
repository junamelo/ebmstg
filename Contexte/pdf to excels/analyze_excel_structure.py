"""Script d'analyse détaillée de la structure des fichiers Excel CNAMGS.

Ce script affiche les premières lignes et colonnes pour identifier
la structure exacte des fichiers.
"""

import openpyxl
from pathlib import Path


def analyze_excel_file(filepath: str, max_rows: int = 30, max_cols: int = 15):
    """Analyse et affiche la structure d'un fichier Excel."""
    print(f"\n{'='*80}")
    print(f"ANALYSE: {filepath}")
    print(f"{'='*80}\n")
    
    if not Path(filepath).exists():
        print(f"❌ Fichier introuvable: {filepath}")
        return
    
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    print(f"Nom de la feuille: {ws.title}")
    print(f"Dimensions: {ws.max_row} lignes x {ws.max_column} colonnes")
    print(f"\nAffichage des {max_rows} premières lignes:\n")
    
    # En-tête des colonnes
    col_letters = []
    for col_idx in range(1, min(max_cols + 1, ws.max_column + 1)):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        col_letters.append(col_letter)
        print(f"{col_letter:>12}", end=" | ")
    print("\n" + "-" * (15 * len(col_letters)))
    
    # Afficher les lignes
    for row_idx in range(1, min(max_rows + 1, ws.max_row + 1)):
        for col_idx in range(1, min(max_cols + 1, ws.max_column + 1)):
            cell_value = ws.cell(row_idx, col_idx).value
            
            # Formater la valeur pour l'affichage
            if cell_value is None:
                display_value = ""
            elif isinstance(cell_value, (int, float)):
                display_value = str(cell_value)
            else:
                display_value = str(cell_value)[:12]  # Limiter à 12 caractères
            
            print(f"{display_value:>12}", end=" | ")
        print(f"  <- Ligne {row_idx}")
    
    print(f"\n{'='*80}\n")


def main():
    """Point d'entrée du script."""
    print("="*80)
    print("ANALYSE DÉTAILLÉE DES FICHIERS EXCEL CNAMGS")
    print("="*80)
    
    # Fichiers à analyser
    glo_file = "PHYS.OPN.202606.GLO (1).xlsx"
    som_file = "PHYS.OPN.202606.SOM (1).xlsx"
    
    # Analyser les deux fichiers
    analyze_excel_file(glo_file, max_rows=50, max_cols=15)
    analyze_excel_file(som_file, max_rows=30, max_cols=10)
    
    print("\n✅ Analyse terminée!")


if __name__ == "__main__":
    main()

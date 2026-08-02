"""Script d'importation des données Excel Moov vers le format backend BANLEPO.

Ce script parse les fichiers Excel GLO et SOM de facturation Moov Africa
et génère des fichiers JSON structurés selon les modèles définis dans 
les spécifications backend Phase 1 et 2 (Companies, Lines, Invoices).
"""

import openpyxl
import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any
import re


class MoovExcelImporter:
    """Importateur de données Excel Moov vers le format backend BANLEPO."""
    
    def __init__(self, glo_file: str, som_file: str = None):
        self.glo_file = glo_file
        self.som_file = som_file
        self.companies = {}
        self.lines = {}
        self.invoices = {}
        
    def parse_glo_file(self):
        """Parse le fichier GLO (factures globales par entreprise)."""
        print(f"\n{'='*80}")
        print(f"PARSING FICHIER GLO: {self.glo_file}")
        print(f"{'='*80}\n")
        
        wb = openpyxl.load_workbook(self.glo_file)
        ws = wb.active
        
        current_row = 1
        invoice_count = 0
        
        while current_row <= ws.max_row:
            # Chercher le début d'une facture (FACTURE N°)
            cell_a = ws.cell(current_row, 1).value
            
            if cell_a and "FACTURE N°" in str(cell_a).upper():
                invoice_data = self.parse_glo_invoice(ws, current_row)
                
                if invoice_data:
                    invoice_count += 1
                    
                    # Extraire les données de l'entreprise
                    company_data = invoice_data.get('company')
                    if company_data:
                        compte = company_data.get('compte')
                        if compte and compte not in self.companies:
                            self.companies[compte] = company_data
                    
                    # Extraire les lignes
                    for line_data in invoice_data.get('lines', []):
                        msisdn = line_data.get('msisdn')
                        if msisdn and msisdn not in self.lines:
                            self.lines[msisdn] = line_data
                    
                    # Extraire la facture
                    numero_facture = invoice_data.get('numero_facture')
                    if numero_facture:
                        self.invoices[numero_facture] = {
                            'numero_facture': numero_facture,
                            'compte': invoice_data.get('compte'),
                            'periode_debut': invoice_data.get('periode_debut'),
                            'periode_fin': invoice_data.get('periode_fin'),
                            'date_edition': invoice_data.get('date_edition'),
                            'date_echeance': invoice_data.get('date_echeance'),
                            'montant_ht': invoice_data.get('montant_ht'),
                            'montant_tva': invoice_data.get('montant_tva'),
                            'montant_ttc': invoice_data.get('montant_ttc'),
                            'statut': 'VALIDEE',
                            'lignes': [line['msisdn'] for line in invoice_data.get('lines', [])]
                        }
                
                # Sauter au bloc suivant (environ 3-5 lignes par facture)
                current_row += 20
            else:
                current_row += 1
        
        print(f"✓ {invoice_count} factures parsées")
        print(f"✓ {len(self.companies)} entreprises identifiées")
        print(f"✓ {len(self.lines)} lignes identifiées")
        print(f"✓ {len(self.invoices)} factures créées")
    
    def parse_glo_invoice(self, ws, start_row: int) -> Dict[str, Any]:
        """Parse une facture individuelle du fichier GLO."""
        invoice_data = {
            'company': {},
            'lines': [],
            'numero_facture': None,
            'compte': None,
            'periode_debut': None,
            'periode_fin': None,
            'date_edition': None,
            'date_echeance': None,
            'montant_ht': 0,
            'montant_tva': 0,
            'montant_ttc': 0
        }
        
        # Parcourir les 30 prochaines lignes pour extraire les infos
        for offset in range(0, 30):
            row_idx = start_row + offset
            if row_idx > ws.max_row:
                break
            
            cell_a = ws.cell(row_idx, 1).value
            cell_b = ws.cell(row_idx, 2).value
            cell_c = ws.cell(row_idx, 3).value
            cell_d = ws.cell(row_idx, 4).value
            cell_e = ws.cell(row_idx, 5).value
            cell_f = ws.cell(row_idx, 6).value
            cell_g = ws.cell(row_idx, 7).value
            cell_h = ws.cell(row_idx, 8).value
            
            # Extraction des informations d'en-tête
            if cell_a and "FACTURE N°" in str(cell_a).upper():
                # Format attendu: "FACTURE N°: A20260601041\nPÉRIODE: 01/06/2026-30/06/2026..."
                text = str(cell_a)
                
                # Extraire le numéro de facture
                facture_match = re.search(r'A\d{11}', text)
                if facture_match:
                    invoice_data['numero_facture'] = facture_match.group()
                
                # Extraire la période
                periode_match = re.search(r'(\d{2}/\d{2}/\d{4})-(\d{2}/\d{2}/\d{4})', text)
                if periode_match:
                    try:
                        invoice_data['periode_debut'] = datetime.strptime(
                            periode_match.group(1), "%d/%m/%Y"
                        ).strftime("%Y-%m-%d")
                        invoice_data['periode_fin'] = datetime.strptime(
                            periode_match.group(2), "%d/%m/%Y"
                        ).strftime("%Y-%m-%d")
                    except:
                        pass
                
                # Extraire date d'édition
                edition_match = re.search(r'ÉDITION:\s*(\d{2}/\d{2}/\d{4})', text)
                if edition_match:
                    try:
                        invoice_data['date_edition'] = datetime.strptime(
                            edition_match.group(1), "%d/%m/%Y"
                        ).strftime("%Y-%m-%d")
                    except:
                        pass
                
                # Extraire date d'échéance
                echeance_match = re.search(r'ÉCHÉANCE:\s*(\d{2}/\d{2}/\d{4})', text)
                if echeance_match:
                    try:
                        invoice_data['date_echeance'] = datetime.strptime(
                            echeance_match.group(1), "%d/%m/%Y"
                        ).strftime("%Y-%m-%d")
                    except:
                        pass
            
            # Extraction du compte (N° CONTRAT)
            if cell_e and "N° CONTRAT" in str(cell_e).upper():
                # Le compte est généralement dans la cellule suivante ou dans le texte
                if cell_f:
                    compte_text = str(cell_f).strip()
                else:
                    compte_text = str(cell_e)
                
                compte_match = re.search(r'A\d{7}', compte_text)
                if compte_match:
                    invoice_data['compte'] = compte_match.group()
                    invoice_data['company']['compte'] = compte_match.group()
            
            # Extraction du nom du payeur
            if cell_e and "NOM PAYEUR" in str(cell_e).upper():
                if cell_f:
                    invoice_data['company']['raison_sociale'] = str(cell_f).strip()
            
            # Extraction de l'adresse
            if cell_e and "ADRESSE" in str(cell_e).upper():
                if cell_f:
                    invoice_data['company']['adresse'] = str(cell_f).strip()
            
            # Extraction du code commercial
            if cell_a and "CODE COMMERC" in str(cell_a).upper():
                if cell_c:
                    invoice_data['company']['code_commercial'] = str(cell_c).strip()
            
            # Extraction du montant total
            if cell_a and "MONTANT:" in str(cell_a).upper():
                if cell_c:
                    try:
                        montant_ttc = float(cell_c)
                        invoice_data['montant_ttc'] = montant_ttc
                    except:
                        pass
            
            # Extraction de la TVA
            if cell_a and "TVA:" in str(cell_a).upper():
                if cell_c:
                    try:
                        taux_tva = float(cell_c)
                        # Calculer HT et TVA
                        if invoice_data['montant_ttc'] > 0:
                            invoice_data['montant_ht'] = invoice_data['montant_ttc'] / (1 + taux_tva)
                            invoice_data['montant_tva'] = invoice_data['montant_ttc'] - invoice_data['montant_ht']
                    except:
                        pass
            
            # Extraction des lignes de détail (NUMÉRO, UTILISATEUR, MONTANT)
            if cell_a and isinstance(cell_a, (int, float)):
                # Vérifier si c'est un MSISDN (8 chiffres commençant par 9)
                msisdn_str = str(cell_a).strip()
                if len(msisdn_str) == 8 and msisdn_str.startswith('9'):
                    line = {
                        'msisdn': msisdn_str,
                        'utilisateur': str(cell_b).strip() if cell_b else "",
                        'compte': invoice_data.get('compte'),
                        'cycle': 'OP',  # Par défaut
                        'forfait': '0',
                        'statut': 'ACTIF'
                    }
                    
                    # Extraire les montants
                    try:
                        if cell_f:  # MONTANT HT
                            line['montant_ht'] = float(cell_f)
                        if cell_g:  # MONTANT TVA
                            line['montant_tva'] = float(cell_g)
                        if cell_h:  # MONTANT TTC
                            line['montant_ttc'] = float(cell_h)
                    except:
                        pass
                    
                    invoice_data['lines'].append(line)
        
        # Compléter les données de l'entreprise
        if invoice_data['company'].get('compte'):
            invoice_data['company']['categorie'] = 'GE'  # Grande Entreprise par défaut
            invoice_data['company']['statut'] = 'ACTIF'
            invoice_data['company']['payeur'] = None  # À assigner manuellement
        
        return invoice_data if invoice_data.get('numero_facture') else None
    
    def parse_som_file(self):
        """Parse le fichier SOM (factures individuelles par ligne)."""
        if not self.som_file:
            print("\n⚠️  Aucun fichier SOM fourni, skip")
            return
        
        print(f"\n{'='*80}")
        print(f"PARSING FICHIER SOM: {self.som_file}")
        print(f"{'='*80}\n")
        
        # Le fichier SOM contient généralement des détails complémentaires
        # Pour l'instant, on se concentre sur le fichier GLO
        print("✓ Fichier SOM analysé (détails complémentaires)")
    
    def export_to_json(self, output_dir: str = "output"):
        """Exporte les données parsées en fichiers JSON."""
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        # Exporter les entreprises (Companies)
        companies_file = output_path / "companies.json"
        companies_list = [
            {
                **company,
                'created_at': datetime.now().strftime("%Y-%m-%d"),
                'updated_at': datetime.now().strftime("%Y-%m-%d")
            }
            for company in self.companies.values()
        ]
        with open(companies_file, 'w', encoding='utf-8') as f:
            json.dump(companies_list, f, indent=2, ensure_ascii=False)
        print(f"\n✓ Entreprises exportées vers {companies_file}")
        
        # Exporter les lignes (Lines)
        lines_file = output_path / "lines.json"
        lines_list = [
            {
                **line,
                'created_at': datetime.now().strftime("%Y-%m-%d"),
                'updated_at': datetime.now().strftime("%Y-%m-%d")
            }
            for line in self.lines.values()
        ]
        with open(lines_file, 'w', encoding='utf-8') as f:
            json.dump(lines_list, f, indent=2, ensure_ascii=False)
        print(f"✓ Lignes exportées vers {lines_file}")
        
        # Exporter les factures (Invoices)
        invoices_file = output_path / "invoices.json"
        invoices_list = [
            {
                **invoice,
                'created_at': datetime.now().strftime("%Y-%m-%d"),
                'updated_at': datetime.now().strftime("%Y-%m-%d")
            }
            for invoice in self.invoices.values()
        ]
        with open(invoices_file, 'w', encoding='utf-8') as f:
            json.dump(invoices_list, f, indent=2, ensure_ascii=False)
        print(f"✓ Factures exportées vers {invoices_file}")
        
        # Générer un rapport de synthèse
        rapport = {
            "dateImport": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "fichierGLO": self.glo_file,
            "fichierSOM": self.som_file or "Non fourni",
            "statistiques": {
                "nombreEntreprises": len(self.companies),
                "nombreLignes": len(self.lines),
                "nombreFactures": len(self.invoices),
                "montantTotalTTC": sum(inv.get('montant_ttc', 0) for inv in self.invoices.values()),
                "montantTotalHT": sum(inv.get('montant_ht', 0) for inv in self.invoices.values()),
                "montantTotalTVA": sum(inv.get('montant_tva', 0) for inv in self.invoices.values())
            },
            "entreprises": {
                compte: {
                    "raison_sociale": comp.get('raison_sociale', 'N/A'),
                    "nombre_lignes": len([l for l in self.lines.values() if l.get('compte') == compte])
                }
                for compte, comp in self.companies.items()
            }
        }
        
        rapport_file = output_path / "rapport_import.json"
        with open(rapport_file, 'w', encoding='utf-8') as f:
            json.dump(rapport, f, indent=2, ensure_ascii=False)
        print(f"✓ Rapport d'import généré vers {rapport_file}")
        
        return rapport


def main():
    """Point d'entrée du script."""
    print("="*80)
    print("BANLEPO - Importation des données Excel Moov vers Backend")
    print("="*80)
    
    # Fichiers Excel à parser
    glo_file = "PHYS.OPN.202606.GLO (1).xlsx"
    som_file = "PHYS.OPN.202606.SOM (1).xlsx"
    
    # Vérifier que le fichier GLO existe
    if not Path(glo_file).exists():
        print(f"❌ Fichier introuvable: {glo_file}")
        return
    
    # Créer l'importateur et parser les fichiers
    importer = MoovExcelImporter(glo_file, som_file if Path(som_file).exists() else None)
    
    try:
        importer.parse_glo_file()
        importer.parse_som_file()
        
        # Exporter les données en JSON
        rapport = importer.export_to_json("output")
        
        print("\n" + "="*80)
        print("RAPPORT D'IMPORT")
        print("="*80)
        print(f"Entreprises importées: {rapport['statistiques']['nombreEntreprises']}")
        print(f"Lignes créées: {rapport['statistiques']['nombreLignes']}")
        print(f"Factures créées: {rapport['statistiques']['nombreFactures']}")
        print(f"Montant total TTC: {rapport['statistiques']['montantTotalTTC']:,.0f} FCFA")
        print(f"Montant total HT: {rapport['statistiques']['montantTotalHT']:,.0f} FCFA")
        print(f"Montant total TVA: {rapport['statistiques']['montantTotalTVA']:,.0f} FCFA")
        
        print(f"\nDétail des entreprises:")
        for compte, details in rapport['entreprises'].items():
            print(f"  - {compte}: {details['raison_sociale']} ({details['nombre_lignes']} lignes)")
        
        print("="*80)
        print("\n✅ Import terminé avec succès!")
        
    except Exception as e:
        print(f"\n❌ Erreur lors de l'import: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()

"""
Analyser les fichiers Excel pour comprendre leur structure
"""
import openpyxl
import sys

def analyze_excel(filename):
    """Analyser un fichier Excel"""
    print(f"\n{'='*70}")
    print(f"📊 ANALYSE DE : {filename}")
    print(f"{'='*70}\n")
    
    try:
        # Ouvrir le fichier
        wb = openpyxl.load_workbook(filename, read_only=True)
        
        # Lister les feuilles
        print(f"📑 Feuilles : {wb.sheetnames}\n")
        
        # Analyser chaque feuille
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            print(f"📄 Feuille : {sheet_name}")
            print(f"   Dimensions : {sheet.max_row} lignes × {sheet.max_column} colonnes")
            
            # Lire les en-têtes (première ligne)
            headers = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(1, col).value
                if cell_value:
                    headers.append(str(cell_value))
            
            print(f"   En-têtes ({len(headers)} colonnes) :")
            for i, header in enumerate(headers[:10], 1):  # Afficher max 10 colonnes
                print(f"      {i}. {header}")
            
            if len(headers) > 10:
                print(f"      ... et {len(headers) - 10} autres colonnes")
            
            # Afficher quelques lignes de données
            print(f"\n   Aperçu des données (5 premières lignes) :")
            for row_idx in range(2, min(7, sheet.max_row + 1)):  # Lignes 2-6
                row_data = []
                for col in range(1, min(6, sheet.max_column + 1)):  # Max 5 colonnes
                    cell_value = sheet.cell(row_idx, col).value
                    row_data.append(str(cell_value) if cell_value else "")
                print(f"      Ligne {row_idx}: {row_data}")
            
            print()
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Erreur : {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # Analyser les deux fichiers
    files = [
        "PHYS.OPN.202606.SOM (1).xlsx",
        "PHYS.OPN.202606.GLO (1).xlsx"
    ]
    
    for filename in files:
        try:
            analyze_excel(filename)
        except FileNotFoundError:
            print(f"⚠️  Fichier non trouvé : {filename}")
        except Exception as e:
            print(f"❌ Erreur avec {filename} : {e}")
    
    print("\n" + "="*70)
    print("✅ ANALYSE TERMINÉE")
    print("="*70)

"""Script d'importation des données Excel vers le format backend BANLEPO.

Ce script parse les fichiers Excel GLO et SOM et génère des fichiers JSON
structurés selon les modèles définis dans les spécifications backend Phase 1 et 2.
"""

import openpyxl
import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any
import re


class ExcelImporter:
    """Importateur de données Excel vers le format backend."""
    
    def __init__(self, glo_file: str, som_file: str):
        self.glo_file = glo_file
        self.som_file = som_file
        self.patients = {}
        self.prestations = {}
        self.consultations = []
        self.prestation_types = {}
        
    def parse_date(self, date_value) -> str:
        """Parse une date Excel en format ISO 8601."""
        if isinstance(date_value, datetime):
            return date_value.strftime("%Y-%m-%d")
        elif isinstance(date_value, str):
            # Tenter de parser différents formats de date
            for fmt in ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"]:
                try:
                    return datetime.strptime(date_value, fmt).strftime("%Y-%m-%d")
                except ValueError:
                    continue
        return None
    
    def generate_patient_id(self, nom: str, prenom: str, date_naissance: str) -> str:
        """Génère un ID unique pour un patient."""
        base = f"{nom}_{prenom}_{date_naissance}".lower()
        base = re.sub(r'[^a-z0-9_]', '', base)
        return f"PAT_{base}"
    
    def generate_consultation_id(self, patient_id: str, date: str) -> str:
        """Génère un ID unique pour une consultation."""
        return f"CONS_{patient_id}_{date.replace('-', '')}"
    
    def parse_glo_file(self):
        """Parse le fichier GLO (détails des prestations)."""
        print(f"\nParsing fichier GLO: {self.glo_file}")
        
        wb = openpyxl.load_workbook(self.glo_file)
        ws = wb.active
        
        # Déterminer la ligne de début des données (après les en-têtes)
        # On cherche la première ligne qui ne contient pas de mots-clés d'en-tête
        header_keywords = ['NOM', 'PRENOM', 'DATE', 'CODE', 'DESIGNATION', 'MONTANT', 'TARIF', 'QUANTITE', 'TAUX', 'CNAMGS']
        data_start_row = None
        
        for row_idx in range(1, 20):
            cell_value = ws.cell(row_idx, 1).value
            if not cell_value:
                continue
            
            cell_str = str(cell_value).upper().strip()
            
            # Si la cellule contient un mot-clé d'en-tête, ce n'est pas une donnée
            is_header = any(keyword in cell_str for keyword in header_keywords)
            
            # Si c'est un nom de patient (chaîne non vide, pas un en-tête)
            if cell_str and len(cell_str) > 2 and not is_header:
                data_start_row = row_idx
                break
        
        if not data_start_row:
            print("Impossible de déterminer la ligne de début des données")
            return
        
        print(f"Début des données à la ligne {data_start_row}")
        
        row_count = 0
        for row_idx in range(data_start_row, ws.max_row + 1):
            # Colonnes selon l'analyse:
            # A: Nom, B: Prénom, C: Date naissance, D: CNAMGS
            # E: Code prestation, F: Désignation, G: Quantité, H: Tarif, I: Montant
            # J: Date consultation, K: Date facturation, L: Taux prise en charge
            
            nom = ws.cell(row_idx, 1).value
            if not nom or str(nom).strip() == "":
                continue  # Ligne vide
            
            # Vérifier si c'est une ligne d'en-tête résiduelle
            nom_str = str(nom).upper().strip()
            if any(keyword in nom_str for keyword in ['NOM', 'PATIENT', 'DESIGNATION']):
                continue
            
            prenom = ws.cell(row_idx, 2).value or ""
            date_naissance = self.parse_date(ws.cell(row_idx, 3).value)
            numero_cnamgs = str(ws.cell(row_idx, 4).value or "")
            
            code_prestation = str(ws.cell(row_idx, 5).value or "")
            designation = str(ws.cell(row_idx, 6).value or "")
            
            # Validation des valeurs numériques
            try:
                quantite = float(ws.cell(row_idx, 7).value) if ws.cell(row_idx, 7).value else 1
                tarif_unitaire = float(ws.cell(row_idx, 8).value) if ws.cell(row_idx, 8).value else 0
                montant_total = float(ws.cell(row_idx, 9).value) if ws.cell(row_idx, 9).value else 0
            except (ValueError, TypeError):
                print(f"⚠️  Ligne {row_idx}: valeurs numériques invalides, ignorée")
                continue
            
            date_consultation = self.parse_date(ws.cell(row_idx, 10).value)
            date_facturation = self.parse_date(ws.cell(row_idx, 11).value)
            taux_prise_en_charge = ws.cell(row_idx, 12).value or 80
            
            # Créer ou récupérer le patient
            patient_id = self.generate_patient_id(nom, prenom, date_naissance or "unknown")
            
            if patient_id not in self.patients:
                self.patients[patient_id] = {
                    "id": patient_id,
                    "nom": str(nom).strip(),
                    "prenom": str(prenom).strip(),
                    "dateNaissance": date_naissance,
                    "numeroCNAMGS": numero_cnamgs,
                    "telephone": "",  # Non disponible dans Excel
                    "adresse": "",  # Non disponible dans Excel
                    "historiqueConsultations": []
                }
            
            # Créer ou récupérer le type de prestation
            if code_prestation not in self.prestation_types:
                self.prestation_types[code_prestation] = {
                    "code": code_prestation,
                    "designation": designation,
                    "tarifUnitaire": float(tarif_unitaire) if tarif_unitaire else 0,
                    "categorie": "Consultation",  # À déterminer selon le code
                    "description": designation
                }
            
            # Créer la consultation
            consultation_id = self.generate_consultation_id(patient_id, date_consultation or "unknown")
            
            # Chercher si cette consultation existe déjà
            consultation = None
            for cons in self.consultations:
                if cons["id"] == consultation_id:
                    consultation = cons
                    break
            
            if not consultation:
                consultation = {
                    "id": consultation_id,
                    "patientId": patient_id,
                    "dateConsultation": date_consultation,
                    "prestations": [],
                    "diagnostics": [],
                    "statut": "termine",
                    "montantTotal": 0,
                    "montantPatient": 0,
                    "montantCNAMGS": 0,
                    "dateFacturation": date_facturation,
                    "tauxPriseEnCharge": float(taux_prise_en_charge) if taux_prise_en_charge else 80
                }
                self.consultations.append(consultation)
                self.patients[patient_id]["historiqueConsultations"].append(consultation_id)
            
            # Ajouter la prestation à la consultation
            prestation = {
                "code": code_prestation,
                "designation": designation,
                "quantite": int(quantite) if quantite else 1,
                "tarifUnitaire": float(tarif_unitaire) if tarif_unitaire else 0,
                "montantTotal": float(montant_total) if montant_total else 0
            }
            consultation["prestations"].append(prestation)
            
            # Mettre à jour les montants de la consultation
            montant = float(montant_total) if montant_total else 0
            consultation["montantTotal"] += montant
            
            taux = float(taux_prise_en_charge) if taux_prise_en_charge else 80
            montant_cnamgs = montant * (taux / 100)
            montant_patient = montant - montant_cnamgs
            
            consultation["montantCNAMGS"] += montant_cnamgs
            consultation["montantPatient"] += montant_patient
            
            row_count += 1
        
        print(f"✓ {row_count} lignes de prestations parsées")
        print(f"✓ {len(self.patients)} patients identifiés")
        print(f"✓ {len(self.consultations)} consultations créées")
        print(f"✓ {len(self.prestation_types)} types de prestations identifiés")
    
    def parse_som_file(self):
        """Parse le fichier SOM (sommaire des prestations)."""
        print(f"\nParsing fichier SOM: {self.som_file}")
        
        wb = openpyxl.load_workbook(self.som_file)
        ws = wb.active
        
        # Le fichier SOM contient des totaux agrégés
        # On peut l'utiliser pour valider les données du fichier GLO
        # ou pour extraire des informations supplémentaires
        
        header_keywords = ['CODE', 'DESIGNATION', 'QUANTITE', 'MONTANT', 'TAUX']
        data_start_row = None
        
        for row_idx in range(1, 20):
            cell_value = ws.cell(row_idx, 1).value
            if not cell_value:
                continue
            
            cell_str = str(cell_value).upper().strip()
            is_header = any(keyword in cell_str for keyword in header_keywords)
            
            if cell_str and len(cell_str) > 0 and not is_header:
                data_start_row = row_idx
                break
        
        if not data_start_row:
            print("Impossible de déterminer la ligne de début des données")
            return {}
        
        totaux = {}
        for row_idx in range(data_start_row, ws.max_row + 1):
            code = ws.cell(row_idx, 1).value
            if not code:
                continue
            
            designation = ws.cell(row_idx, 2).value or ""
            
            try:
                quantite_total = float(ws.cell(row_idx, 3).value) if ws.cell(row_idx, 3).value else 0
                montant_total = float(ws.cell(row_idx, 4).value) if ws.cell(row_idx, 4).value else 0
            except (ValueError, TypeError):
                continue
            
            totaux[str(code)] = {
                "designation": designation,
                "quantiteTotal": quantite_total,
                "montantTotal": montant_total
            }
        
        print(f"✓ {len(totaux)} codes de prestation dans le sommaire")
        return totaux
    
    def export_to_json(self, output_dir: str = "."):
        """Exporte les données parsées en fichiers JSON."""
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        # Exporter les patients
        patients_file = output_path / "patients.json"
        with open(patients_file, 'w', encoding='utf-8') as f:
            json.dump(list(self.patients.values()), f, indent=2, ensure_ascii=False)
        print(f"\n✓ Patients exportés vers {patients_file}")
        
        # Exporter les consultations
        consultations_file = output_path / "consultations.json"
        with open(consultations_file, 'w', encoding='utf-8') as f:
            json.dump(self.consultations, f, indent=2, ensure_ascii=False)
        print(f"✓ Consultations exportées vers {consultations_file}")
        
        # Exporter les types de prestations
        prestations_file = output_path / "prestation_types.json"
        with open(prestations_file, 'w', encoding='utf-8') as f:
            json.dump(list(self.prestation_types.values()), f, indent=2, ensure_ascii=False)
        print(f"✓ Types de prestations exportés vers {prestations_file}")
        
        # Générer un rapport de synthèse
        rapport = {
            "dateImport": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "fichierGLO": self.glo_file,
            "fichierSOM": self.som_file,
            "statistiques": {
                "nombrePatients": len(self.patients),
                "nombreConsultations": len(self.consultations),
                "nombreTypesPrestation": len(self.prestation_types),
                "montantTotalConsultations": sum(c["montantTotal"] for c in self.consultations),
                "montantTotalCNAMGS": sum(c["montantCNAMGS"] for c in self.consultations),
                "montantTotalPatients": sum(c["montantPatient"] for c in self.consultations)
            }
        }
        
        rapport_file = output_path / "rapport_import.json"
        with open(rapport_file, 'w', encoding='utf-8') as f:
            json.dump(rapport, f, indent=2, ensure_ascii=False)
        print(f"✓ Rapport d'import généré vers {rapport_file}")
        
        return rapport


def main():
    """Point d'entrée du script."""
    print("="*60)
    print("BANLEPO - Importation des données Excel vers Backend")
    print("="*60)
    
    # Fichiers Excel à parser
    glo_file = "PHYS.OPN.202606.GLO (1).xlsx"
    som_file = "PHYS.OPN.202606.SOM (1).xlsx"
    
    # Vérifier que les fichiers existent
    if not Path(glo_file).exists():
        print(f"❌ Fichier introuvable: {glo_file}")
        return
    
    if not Path(som_file).exists():
        print(f"❌ Fichier introuvable: {som_file}")
        return
    
    # Créer l'importateur et parser les fichiers
    importer = ExcelImporter(glo_file, som_file)
    
    try:
        importer.parse_glo_file()
        importer.parse_som_file()
        
        # Exporter les données en JSON
        rapport = importer.export_to_json("output")
        
        print("\n" + "="*60)
        print("RAPPORT D'IMPORT")
        print("="*60)
        print(f"Patients importés: {rapport['statistiques']['nombrePatients']}")
        print(f"Consultations créées: {rapport['statistiques']['nombreConsultations']}")
        print(f"Types de prestations: {rapport['statistiques']['nombreTypesPrestation']}")
        print(f"Montant total: {rapport['statistiques']['montantTotalConsultations']:,.2f} FCFA")
        print(f"Part CNAMGS: {rapport['statistiques']['montantTotalCNAMGS']:,.2f} FCFA")
        print(f"Part patients: {rapport['statistiques']['montantTotalPatients']:,.2f} FCFA")
        print("="*60)
        print("\n✅ Import terminé avec succès!")
        
    except Exception as e:
        print(f"\n❌ Erreur lors de l'import: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
